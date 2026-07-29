#!/usr/bin/env python
"""Validate internal consistency of the public source-data workbook.

All quantitative results are recomputed from workbook rows. The validator does
not use manuscript result constants as pass/fail targets; it cross-checks
independent workbook sheets and the rounded summary table.
"""
from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from statistics import mean, pstdev

from openpyxl import load_workbook


def rows(path: Path, sheet: str) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    data = list(ws.iter_rows(values_only=True))
    if not data:
        raise AssertionError(f"Empty sheet: {sheet}")
    headers = [str(x).strip() for x in data[0]]
    return [
        {headers[i]: row[i] for i in range(len(headers))}
        for row in data[1:]
        if any(value is not None for value in row)
    ]


def pearson(x, y) -> float:
    x = [float(v) for v in x]
    y = [float(v) for v in y]
    mx, my = mean(x), mean(y)
    numerator = sum((a - mx) * (b - my) for a, b in zip(x, y))
    denominator = math.sqrt(
        sum((a - mx) ** 2 for a in x) * sum((b - my) ** 2 for b in y)
    )
    return numerator / denominator


def ranks(values) -> list[float]:
    indexed = sorted((float(value), index) for index, value in enumerate(values))
    out = [0.0] * len(indexed)
    i = 0
    while i < len(indexed):
        j = i + 1
        while j < len(indexed) and indexed[j][0] == indexed[i][0]:
            j += 1
        rank = (i + 1 + j) / 2
        for k in range(i, j):
            out[indexed[k][1]] = rank
        i = j
    return out


def spearman(x, y) -> float:
    return pearson(ranks(x), ranks(y))


def close(a, b, tol=1e-8, label="value") -> None:
    if abs(float(a) - float(b)) > tol:
        raise AssertionError(f"{label}: {a} != {b} within {tol}")


def rounded_close(computed, reported, digits=4, label="reported value") -> None:
    close(round(float(computed), digits), float(reported), 0.5 * 10 ** (-digits), label)


def parse_fraction(value: object) -> tuple[int, int]:
    match = re.match(r"\s*(\d+)\s*/\s*(\d+)", str(value))
    if not match:
        raise AssertionError(f"Cannot parse fraction: {value!r}")
    return int(match.group(1)), int(match.group(2))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--workbook", default="source_data/Supplementary_Data_S1.xlsx"
    )
    args = parser.parse_args()
    path = Path(args.workbook)

    inventory = rows(path, "Genome inventory")
    accessions = [str(row["Assembly accession"]) for row in inventory]
    if len(accessions) != len(set(accessions)):
        raise AssertionError("Genome inventory contains duplicate accessions")
    if any(accession.startswith("GCF_000149035.1") for accession in accessions):
        raise AssertionError("Redundant RefSeq copy remains in the genome inventory")

    main_table = {row["Metric"]: row for row in rows(path, "Main Table 2 data")}

    pairs = rows(path, "WGS distance pairs")
    expected_pairs = len(inventory) * (len(inventory) - 1) // 2
    if len(pairs) != expected_pairs:
        raise AssertionError(f"Pair table has {len(pairs)} rows; expected {expected_pairs}")
    rapd = [row["RAPD band distance"] for row in pairs]
    mash = [row["Mash distance"] for row in pairs]
    pearson_r = pearson(rapd, mash)
    spearman_r = spearman(rapd, mash)
    rounded_close(
        pearson_r,
        main_table["RAPD-vs-Mash Pearson r"]["Value"],
        4,
        "Main Table 2 Pearson r",
    )
    rounded_close(
        spearman_r,
        main_table["RAPD-vs-Mash Spearman r"]["Value"],
        4,
        "Main Table 2 Spearman r",
    )

    method_rows = rows(path, "WGS method correlations")
    method_main = [
        row
        for row in method_rows
        if row["Comparison"] == "RAPD distance vs Mash distance"
        and row["Subset"] == "All pairs"
    ][0]
    close(pearson_r, method_main["Pearson r"], 5e-8, "method-summary Pearson r")
    close(spearman_r, method_main["Spearman r"], 5e-8, "method-summary Spearman r")

    loo = rows(path, "WGS leave-one-out")
    if len(loo) != len(inventory):
        raise AssertionError("Leave-one-out sheet does not contain one row per assembly")
    loo_spearman = []
    for row in loo:
        omitted = str(row["Assembly omitted"]).split()[0]
        subset = [
            pair
            for pair in pairs
            if not str(pair["Assembly A"]).startswith(omitted)
            and not str(pair["Assembly B"]).startswith(omitted)
        ]
        expected_subset = (len(inventory) - 1) * (len(inventory) - 2) // 2
        if len(subset) != expected_subset:
            raise AssertionError(f"Unexpected leave-one-out pair count for {omitted}")
        p = pearson(
            [item["RAPD band distance"] for item in subset],
            [item["Mash distance"] for item in subset],
        )
        s = spearman(
            [item["RAPD band distance"] for item in subset],
            [item["Mash distance"] for item in subset],
        )
        close(p, row["Pearson r"], 5e-8, f"LOO Pearson r for {omitted}")
        close(s, row["Spearman r"], 5e-8, f"LOO Spearman r for {omitted}")
        loo_spearman.append(s)
    minimum_loo_spearman = min(loo_spearman)
    rounded_close(
        minimum_loo_spearman,
        main_table["Minimum leave-one-assembly-out Spearman r"]["Value"],
        4,
        "Main Table 2 minimum LOO Spearman r",
    )

    context = rows(path, "Genome context summary")
    overall = {
        row["Metric"]: row
        for row in context
        if row["Grouping variable"] == "Overall" and row["Group value"] == "All"
    }
    iterations = int(next(iter(overall.values()))["Random iterations"])

    primer_sets = rows(path, "Random primer sets")
    historical = [row for row in primer_sets if row["Set type"] == "Historical"]
    random_sets = [row for row in primer_sets if row["Set type"] == "GC-matched random"]
    if len(historical) != 1 or not random_sets:
        raise AssertionError("Primer-set classification is incomplete")
    historical = historical[0]
    amplicons = int(historical["Total amplicons"])
    matched_random_intervals = amplicons * iterations
    if int(main_table["Observed RAPD amplicons"]["Value"]) != amplicons:
        raise AssertionError("Main Table 2 amplicon count does not match primer-set rows")
    if int(main_table["Matched random intervals"]["Value"]) != matched_random_intervals:
        raise AssertionError("Main Table 2 random-interval count is inconsistent")
    if int(main_table["GC-matched random primer sets retained"]["Value"]) != len(random_sets):
        raise AssertionError("Main Table 2 random-primer count is inconsistent")

    null_rows = {row["Metric"]: row for row in rows(path, "Random primer null")}
    metric_map = {
        "Total amplicons": "Total amplicons",
        "Band-size bins": "Band-size bins",
        "Mean bands per assembly": "Mean bands per assembly",
        "Fixed-core band fraction": "Fixed-core band fraction",
        "Polymorphic band fraction": "Polymorphic band fraction",
        "Rare/private band fraction": "Rare/private band fraction",
        "Informative bands": "Informative bands",
        "Pearson r (non-duplicate pairs)": "Pearson r (non-duplicate pairs)",
        "Spearman r (non-duplicate pairs)": "Spearman r (non-duplicate pairs)",
        "Minimum leave-one-out Pearson r": "Minimum leave-one-out Pearson r",
        "Minimum leave-one-out Spearman r": "Minimum leave-one-out Spearman r",
        "Fraction within 5 kb of contig end": "Fraction within 5 kb of contig end",
        "Fraction within 10 kb of contig end": "Fraction within 10 kb of contig end",
        "Fraction within 25 kb of contig end": "Fraction within 25 kb of contig end",
        "Mean GC fraction": "Mean GC fraction",
        "Mean N fraction": "Mean N fraction",
        "Mean Shannon entropy": "Mean Shannon entropy",
        "Mean maximum homopolymer run": "Mean maximum homopolymer run",
    }
    for primer_column, null_metric in metric_map.items():
        if null_metric not in null_rows:
            continue
        target = null_rows[null_metric]
        close(
            historical[primer_column],
            target["Historical primer value"],
            5e-8,
            f"historical {null_metric}",
        )
        close(
            mean(float(row[primer_column]) for row in random_sets),
            target["Random mean"],
            5e-8,
            f"random mean {null_metric}",
        )
        if int(target["Random primer sets"]) != len(random_sets):
            raise AssertionError(f"Random-set count mismatch for {null_metric}")

    guthrie = rows(path, "Guthrie summary agreement")
    global_rows = [row for row in guthrie if row["Denominator"] == "Global denominator"]
    differences = [
        float(row["Digitized shared-locus (%)"])
        - float(row["Published shared-locus (%)"])
        for row in global_rows
    ]
    rmse = math.sqrt(mean(value * value for value in differences))
    composite_loci = max(int(row["Composite loci"]) for row in global_rows)
    if int(main_table["Composite schematic loci"]["Value"]) != composite_loci:
        raise AssertionError("Composite-locus count is inconsistent")
    rounded_close(
        rmse,
        main_table["Table 3 RMSE"]["Value"],
        3,
        "Main Table 2 RMSE",
    )

    tolerance_rows = rows(path, "Guthrie bridge tolerance")
    selected = min(
        tolerance_rows,
        key=lambda row: abs(float(row["Band-size tolerance (kb)"]) - 0.05),
    )
    legacy_total = int(selected["Legacy bands"])
    legacy_matched = int(selected["Legacy bands with modern match"])
    reported_matched, reported_total = parse_fraction(
        main_table["Legacy bins matched to modern band-size space"]["Value"]
    )
    if (legacy_matched, legacy_total) != (reported_matched, reported_total):
        raise AssertionError("Any-bridge summary is inconsistent")

    bridge_classes = rows(path, "Guthrie bridge classes")
    class_counts = {
        label: sum(row["Bridge class"] == label for row in bridge_classes)
        for label in ("Polymorphic bridge", "Core-only bridge", "No modern match")
    }
    if sum(class_counts.values()) != legacy_total:
        raise AssertionError("Bridge classes do not partition all legacy bands")
    reported_poly, reported_poly_total = parse_fraction(
        main_table["Legacy bins matched to modern polymorphic windows"]["Value"]
    )
    if reported_poly_total != legacy_total or reported_poly != class_counts["Polymorphic bridge"]:
        raise AssertionError("Polymorphic-bridge summary is inconsistent")
    if class_counts["Polymorphic bridge"] + class_counts["Core-only bridge"] != legacy_matched:
        raise AssertionError("Bridge-class counts do not match the selected tolerance summary")

    segments = rows(path, "Denoyes detected segments")
    lanes = rows(path, "Denoyes lane audit")
    retained_lanes = sum(row["Lane status"] == "Retained sample lane" for row in lanes)
    manual_lanes = sum(row["Lane status"] == "Manual-check sample lane" for row in lanes)
    if int(main_table["Detected true-gel band segments"]["Value"]) != len(segments):
        raise AssertionError("Denoyes segment count is inconsistent")
    reported_retained, reported_total_lanes = parse_fraction(
        main_table["Nonzero lanes retained"]["Value"]
    )
    if reported_retained != retained_lanes or reported_total_lanes != retained_lanes + manual_lanes:
        raise AssertionError("Denoyes lane summary is inconsistent")

    print("Source-data validation passed")
    print(f"- Assemblies: {len(inventory)}")
    print(f"- RAPD/Mash pairwise comparisons: {len(pairs)}")
    print(f"- RAPD-vs-Mash Pearson r: {pearson_r:.6f}")
    print(f"- RAPD-vs-Mash Spearman r: {spearman_r:.6f}")
    print(f"- Minimum leave-one-assembly-out Spearman r: {minimum_loo_spearman:.6f}")
    print(f"- Gene overlap observed/random: {float(overall['Fraction overlapping annotated genes']['Observed value']):.6f} / {float(overall['Fraction overlapping annotated genes']['Random mean']):.6f}")
    print(f"- Mean GC observed/random: {float(overall['Mean GC fraction']['Observed value']):.6f} / {float(overall['Mean GC fraction']['Random mean']):.6f}")
    print(f"- Historical amplicons: {amplicons}")
    print(f"- Accepted random primer sets: {len(random_sets)}")
    print(f"- Historical/random non-duplicate Spearman r: {float(historical['Spearman r (non-duplicate pairs)']):.6f} / {mean(float(row['Spearman r (non-duplicate pairs)']) for row in random_sets):.6f}")
    print(f"- Guthrie RMSE: {rmse:.6f}")
    print(f"- Legacy bridge: {legacy_matched}/{legacy_total}; classes={class_counts}")
    print(f"- Denoyes segments: {len(segments)}; retained/manual-check lanes: {retained_lanes}/{manual_lanes}")


if __name__ == "__main__":
    main()
